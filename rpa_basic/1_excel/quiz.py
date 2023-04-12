from openpyxl import Workbook, load_workbook

wb = Workbook()
ws = wb.active

arr = [["학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"],
       [1, 10, 8, 5, 14, 26, 12],
       [2, 7, 3, 7, 15, 24, 18],
       [3, 9, 5, 8, 8, 12, 4],
       [4, 7, 8, 7, 17, 21, 18],
       [5, 7, 8, 7, 16, 25, 15],
       [6, 3, 5, 8, 8, 17, 0],
       [7, 4, 9, 10, 16, 27, 18],
       [8, 6, 6, 6, 15, 19, 17],
       [9, 10, 10, 9, 19, 30, 19],
       [10, 9, 8, 8, 20, 25, 20]]

for x in range(0, (len(arr))):
    ws.append(arr[x])

for x in range(1, 11):
    ws[f"D{x+1}"].value= 10

ws["H1"] = "총점"
ws["I1"] = "성적"

for x in range(1, 11):
    ws[f"H{x+1}"] = sum(arr[x][1:])-arr[x][3] + 10
    if ws[f"B{x+1}"].value < 5:
        ws[f"I{x+1}"].value = "F"
        continue

    if ws[f"H{x+1}"].value >= 90:
        ws[f"I{x+1}"].value = "A"
    elif ws[f"H{x+1}"].value >= 80:
        ws[f"I{x+1}"].value = "B"
    elif ws[f"H{x+1}"].value >= 70:
        ws[f"I{x+1}"].value = "C"
    else:
        ws[f"I{x+1}"].value = "D"

wb.save("scores.xlsx")
