from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

for row in ws.iter_rows(min_row=2):
    if int(row[1].value) > 80:
        print(row[0].value, "번 학생은 영어 천재")
